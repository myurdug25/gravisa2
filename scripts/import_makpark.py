#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Makpark.xlsx → data/makineler_admin.json (admin panel şeması).

Gereksinim: pip install openpyxl

Kullanım:
  python scripts/import_makpark.py
  python scripts/import_makpark.py "C:\\Yol\\Makpark.xlsx"

Varsayılan xlsx yolu: proje kökünde Makpark.xlsx

Makpark.xlsx: İki satır başlık — 1. satırda TYPE = kategori, TİP = model kodu; veri 3. satırdan başlar.

Diğer Excel: Sütun başlıkları (Türkçe/İngilizce) otomatik eşlenir:
  No, TYPE/Tip/kategori, Firma, Tip/Model, Model Yılı, Güç, Kapasite, Şasi, Motor, Stok

Kategori (tip): TYPE sütunundan okunur; YERALTI → YER ALTI gibi tutarlı yazım uygulanır.
Eski JSON’daki kategori adlarına otomatik eşleme yapılmaz (yanlış sınıflandırma önlenir).

Görsel: images/makineler/makine_{envanter no}.(webp|jpg|png|jfif) klasörde ne varsa envanter
numarasına (no) göre atanır.

Mevcut JSON’u yeniden taramak için:
  python scripts/import_makpark.py --assign-images-only
"""

from __future__ import annotations

import json
import re
import shutil
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "Makpark.xlsx"
OUT_JSON = ROOT / "data" / "makineler_admin.json"
BACKUP_JSON = ROOT / "data" / "makineler_admin.json.bak"
IMG_DIR = ROOT / "images" / "makineler"
EXISTING_JSON = ROOT / "data" / "makineler_admin.json"
# Aynı numara için birden fazla uzantı varsa öncelik (daha küçük = önce)
_IMG_EXT_PRIORITY = {".webp": 0, ".jpg": 1, ".jpeg": 2, ".png": 3, ".jfif": 4}


def fold_tr(s: str) -> str:
    s = (s or "").strip().lower()
    for a, b in (
        ("ğ", "g"),
        ("ü", "u"),
        ("ş", "s"),
        ("ı", "i"),
        ("i̇", "i"),
        ("ö", "o"),
        ("ç", "c"),
    ):
        s = s.replace(a, b)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", s)


def normalize_header(cell) -> str:
    s = str(cell if cell is not None else "").strip().lower()
    s = s.replace("ı", "i")
    return re.sub(r"\s+", " ", s)


# Excel başlık → iç alan adı
HEADER_ALIASES: dict[str, str] = {}


def _register_header_aliases(keys: tuple, field: str) -> None:
    for k in keys:
        nh = normalize_header(k)
        HEADER_ALIASES[nh] = field
        compact = nh.replace(" ", "").replace("/", "")
        HEADER_ALIASES[compact] = field


for keys, field in (
    (("no", "sıra", "sira", "#", "numara", "satir"), "no"),
    # Makpark.xlsx: kategori sütunu TYPE; TİP = model kodu (ayrı sütun)
    (("type", "category", "kategori", "makine tipi", "grup"), "tip"),
    (("tip", "tür", "tur", "makine tipi"), "tip"),
    (("firma", "marka", "üretici", "uretici"), "firma"),
    (("tipmodel", "tip / model", "tip/model", "model", "tip model"), "tipModel"),
    (("modelyili", "model yılı", "model yili", "yil", "yıl"), "modelYil"),
    (("güç", "guc", "power"), "guc_raw"),
    (("kapasite", "kap."), "kapasite"),
    (("şasiserino", "sase seri no", "şasi seri no", "sasi seri no", "şasi", "sasi"), "saseSeriNo"),
    (("motorserino", "motor seri no"), "motorSeriNo"),
    (("motormarka", "motor markası", "motor markasi"), "motorMarka"),
    (("motortip", "motor tipi"), "motorTip"),
    (("stok", "stokta"), "stok_raw"),
):
    _register_header_aliases(keys, field)


def map_header_row(row) -> dict[int, str]:
    """Sütun indeksi → alan adı."""
    out: dict[int, str] = {}
    for i, cell in enumerate(row):
        if cell is None or str(cell).strip() == "":
            continue
        key = normalize_header(str(cell))
        key_compact = key.replace(" ", "")
        field = HEADER_ALIASES.get(key) or HEADER_ALIASES.get(key_compact)
        if field:
            out[i] = field
    return out


def is_makpark_double_header(row0: tuple) -> bool:
    """Makpark şablonu: 1. satırda 2. sütun başlığı TYPE."""
    if not row0 or len(row0) < 2:
        return False
    return fold_tr(str(row0[1] or "")) == "type"


def makpark_colmap() -> dict[int, str]:
    """
    Makpark.xlsx sütunları (0 tabanlı):
    No | TYPE | FIRMA | TİP(model) | MODEL(yıl) | GÜÇ | KAPASİTE | ŞASE | MOTOR SN | MOTOR Marka | Motor Tip
    """
    return {
        0: "no",
        1: "tip",
        2: "firma",
        3: "tipModel",
        4: "modelYil",
        5: "guc_raw",
        6: "kapasite",
        7: "saseSeriNo",
        8: "motorSeriNo",
        9: "motorMarka",
        10: "motorTip",
    }


def cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        val = int(val)
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def parse_guc(raw: str) -> tuple[str, str]:
    t = (raw or "").strip()
    if not t:
        return "", ""
    m = re.search(r"([\d.,]+)\s*(kW|HP|kVA)\b", t, re.I)
    if m:
        num = m.group(1).replace(",", ".")
        unit = m.group(2).upper()
        if unit == "KW":
            unit = "kW"
        elif unit == "KVA":
            unit = "kVA"
        elif unit == "HP":
            unit = "HP"
        return num, unit
    return t, ""


def parse_stok(raw) -> bool:
    if raw is None or raw == "":
        return True
    s = str(raw).strip().lower()
    if s in ("0", "hayır", "hayir", "yok", "no", "false", "x"):
        return False
    return True


def clean_display_cell(s: str) -> str:
    """Fazla boşluk ve sondaki gereksiz virgül."""
    s = re.sub(r"\s+", " ", (s or "").strip())
    s = re.sub(r"\s*,\s*$", "", s)
    return s


def normalize_category_label(s: str) -> str:
    """
    TYPE sütunu → sitede tutarlı kategori metni.
    Excel’de sık yazım: YERALTI (tek kelime) — slug için YER ALTI ile aynı olmalı (yer-alti-...).
    """
    s = clean_display_cell(s)
    if not s:
        return s
    # Başta YERALTI / Yeraltı → YER ALTI
    if re.match(r"(?i)yeraltı\b", s) or re.match(r"(?i)yeralti\b", s):
        s = re.sub(r"(?i)^yeraltı\b", "YER ALTI", s)
        s = re.sub(r"(?i)^yeralti\b", "YER ALTI", s)
    return s


def machine_fingerprint(m: dict) -> str:
    """Tam aynı teknik satırın iki kez yazılmasını ayırt etmek için."""
    parts = [
        fold_tr(str(m.get("tip") or "")),
        fold_tr(str(m.get("firma") or "")),
        fold_tr(str(m.get("tipModel") or "")),
        fold_tr(str(m.get("modelYil") or "")),
        fold_tr(str(m.get("saseSeriNo") or "")),
        fold_tr(str(m.get("motorSeriNo") or "")),
    ]
    return "|".join(parts)


def resolve_tip(raw: str) -> str:
    """
    Kategori yalnızca Excel TYPE sütunundan (normalize edilmiş).
    Eski JSON’daki kısa/yanlış kategori adlarına geri çekilmez (BETON ≠ BETON YAĞ PÜSKÜRTME).
    """
    raw = normalize_category_label((raw or "").strip())
    return raw if raw else "DİĞER"


def scan_machine_images(img_dir: Path) -> dict[int, str]:
    """makine_{n}.ext dosyalarını tarar; numara -> web yolu (kaliteli uzantı önceliği)."""
    if not img_dir.is_dir():
        return {}
    pat = re.compile(r"^makine_(\d+)\.([a-zA-Z0-9]+)$")
    found: dict[int, list[tuple[int, str]]] = {}
    for f in img_dir.iterdir():
        if not f.is_file():
            continue
        m = pat.match(f.name)
        if not m:
            continue
        num = int(m.group(1))
        ext = "." + m.group(2).lower()
        if ext not in _IMG_EXT_PRIORITY:
            continue
        pri = _IMG_EXT_PRIORITY[ext]
        rel = f"images/makineler/{f.name}"
        found.setdefault(num, []).append((pri, rel))
    out: dict[int, str] = {}
    for num, pairs in found.items():
        pairs.sort(key=lambda x: x[0])
        out[num] = pairs[0][1]
    return out


def image_for_inventory_no(no: str, index: dict[int, str]) -> str:
    """Excel/envanter sıra numarası (no) ile görsel yolu."""
    s = str(no).strip()
    if not s.isdigit():
        return ""
    return index.get(int(s), "")


def row_to_machine(
    row: tuple,
    colmap: dict[int, str],
    machine_id: int,
) -> dict | None:
    cells: dict[str, str] = {}
    for idx, val in enumerate(row):
        if idx not in colmap:
            continue
        field = colmap[idx]
        cells[field] = cell_to_str(val)

    tip = resolve_tip(cells.get("tip", ""))
    firma = clean_display_cell(cells.get("firma", ""))
    tip_model = clean_display_cell(cells.get("tipModel", ""))
    if not tip_model and not firma and not cells.get("no"):
        # Muhtemelen boş satır
        if not any(str(v).strip() for v in cells.values()):
            return None

    guc_val, guc_birim = parse_guc(cells.get("guc_raw", ""))

    no = cells.get("no", "") or str(machine_id)
    stok = parse_stok(cells.get("stok_raw"))

    return {
        "id": machine_id,
        "no": no,
        "tip": tip,
        "firma": firma,
        "tipModel": tip_model,
        "modelYil": clean_display_cell(cells.get("modelYil", "")),
        "guc": guc_val,
        "gucBirim": guc_birim,
        "kapasite": clean_display_cell(cells.get("kapasite", "")),
        "saseSeriNo": clean_display_cell(cells.get("saseSeriNo", "")),
        "motorSeriNo": clean_display_cell(cells.get("motorSeriNo", "")),
        "motorMarka": clean_display_cell(cells.get("motorMarka", "")),
        "motorTip": clean_display_cell(cells.get("motorTip", "")),
        "stok": stok,
        "img": "",
    }


def _is_makpark_data_row(row: tuple) -> bool:
    if not row or row[0] is None:
        return False
    if isinstance(row[0], (int, float)) and not isinstance(row[0], bool):
        return True
    s = str(row[0]).strip()
    return s.isdigit()


def main() -> int:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        print("openpyxl gerekli: pip install openpyxl", file=sys.stderr)
        return 1

    xlsx = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        print(f"Dosya bulunamadı: {xlsx}", file=sys.stderr)
        print("Makpark.xlsx dosyasını proje köküne koyun veya tam yolu argüman olarak verin.", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        print("Excel boş.", file=sys.stderr)
        return 1

    use_makpark = is_makpark_double_header(tuple(all_rows[0]))
    if use_makpark:
        colmap = makpark_colmap()
        data_rows = all_rows[2:]
    else:
        colmap = map_header_row(tuple(all_rows[0]))
        data_rows = all_rows[1:]

    vals = set(colmap.values())
    if "tipModel" not in vals and "tip" not in vals:
        print(
            "Başlık satırında en az kategori (TYPE/Tip) veya Tip/Model sütunu gerekli. "
            f"Eşlenen sütunlar: {sorted(vals)}",
            file=sys.stderr,
        )
        return 1

    img_index = scan_machine_images(IMG_DIR)
    machines: list[dict] = []
    seen_fp: set[str] = set()
    seen_no: set[str] = set()
    skipped_dup = 0
    for row in data_rows:
        if use_makpark and not _is_makpark_data_row(tuple(row or ())):
            continue
        m = row_to_machine(tuple(row or ()), colmap, len(machines) + 1)
        if m is None:
            continue
        fp = machine_fingerprint(m)
        if fp in seen_fp:
            skipped_dup += 1
            continue
        no_s = str(m.get("no", "")).strip()
        if no_s and no_s in seen_no:
            skipped_dup += 1
            continue
        seen_fp.add(fp)
        if no_s:
            seen_no.add(no_s)
        m["img"] = image_for_inventory_no(str(m.get("no", m["id"])), img_index)
        machines.append(m)

    if not machines:
        print("Hiç makine satırı okunamadı.", file=sys.stderr)
        return 1

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    if OUT_JSON.is_file():
        shutil.copy2(OUT_JSON, BACKUP_JSON)

    OUT_JSON.write_text(
        json.dumps(machines, ensure_ascii=False, indent=4),
        encoding="utf-8",
    )
    if skipped_dup:
        print(f"Uyarı: {skipped_dup} tekrarlayan Excel satırı atlandı (aynı no veya aynı teknik içerik).", file=sys.stderr)
    print(f"{len(machines)} makine yazıldı: {OUT_JSON}")
    if BACKUP_JSON.is_file():
        print(f"Önceki liste yedek: {BACKUP_JSON}")
    return 0


def assign_images_only() -> int:
    """makineler_admin.json içindeki tüm kayıtlara images/makineler eşlemesini yeniler."""
    if not OUT_JSON.is_file():
        print(f"Bulunamadı: {OUT_JSON}", file=sys.stderr)
        return 1
    img_index = scan_machine_images(IMG_DIR)
    data = json.loads(OUT_JSON.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        print("Geçersiz JSON", file=sys.stderr)
        return 1
    filled = 0
    for m in data:
        if not isinstance(m, dict):
            continue
        no = str(m.get("no", m.get("id", ""))).strip()
        m["img"] = image_for_inventory_no(no, img_index)
        if m["img"]:
            filled += 1
    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=4), encoding="utf-8")
    print(f"{len(data)} kayıt, {filled} görsel atanmış (klasör: {IMG_DIR})")
    return 0


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--assign-images-only":
        raise SystemExit(assign_images_only())
    raise SystemExit(main())
